"""出料表 / 切割单导出 —— Excel(.xlsx,openpyxl)+ 打印友好 HTML(供浏览器导出 PDF)。

Excel 是制造现场真正要的交付物(可打印、ERP 可导入)。
PDF 走「打印友好 HTML → 浏览器打印为 PDF」,规避后端中文字体坑。
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEAD = Font(bold=True, color="FFFFFF")
_HEADFILL = PatternFill("solid", fgColor="4E79A7")
_REMFILL = PatternFill("solid", fgColor="E2EFDA")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEAD
        cell.fill = _HEADFILL
        cell.alignment = _CENTER


def _autofit(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _counts_text(part_counts: dict) -> str:
    return ", ".join(f"{k}×{v}" for k, v in part_counts.items())


def sheet_xlsx(result: dict) -> bytes:
    wb = Workbook()
    spec = result.get("spec", {})
    acc = result.get("accounting", {})

    # 汇总
    ws = wb.active
    ws.title = "汇总"
    rows = [
        ("项目", "数值"),
        ("材料牌号", spec.get("material", "-")),
        ("板厚(mm)", spec.get("thickness", "-")),
        ("总利用率", f"{result['overall_utilization'] * 100:.1f}%"),
        ("用新板(张)", result["sheets_used"]),
        ("用余料(张)", result.get("remnants_used", 0)),
        ("已排零件(件)", result["total_placed"]),
        ("缺口(件)", result["total_shortfall"]),
    ]
    if acc:
        rows += [
            ("每平米重量(kg/m²)", acc.get("weight_kg_per_m2", "-")),
            ("净产品重量(kg)", acc.get("product_weight_kg", "-")),
            ("投入新板重量(kg)", acc.get("input_weight_kg", "-")),
            ("可回收余料(kg)", acc.get("remnant_weight_kg", "-")),
        ]
    for r in rows:
        ws.append(r)
    _style_header(ws, 1, 2)
    _autofit(ws, [22, 22])

    # 出料表
    ws2 = wb.create_sheet("出料表")
    ws2.append(["板料编码", "类型", "零件号 × 数量", "利用率"])
    _style_header(ws2, 1, 4)
    for s in result["sheets"]:
        ws2.append([s["sheet_code"], "余料" if s["is_remnant"] else "新板",
                    _counts_text(s["part_counts"]), f"{s['utilization'] * 100:.1f}%"])
        if s["is_remnant"]:
            for c in range(1, 5):
                ws2.cell(row=ws2.max_row, column=c).fill = _REMFILL
    _autofit(ws2, [16, 8, 50, 10])

    # 零件汇总/缺口
    ws3 = wb.create_sheet("零件汇总")
    ws3.append(["零件号", "需求", "已排", "缺口"])
    _style_header(ws3, 1, 4)
    for p in result["per_part"]:
        ws3.append([p["part_id"], p["demand"], p["placed"], p["shortfall"]])
    _autofit(ws3, [14, 10, 10, 10])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def bar_xlsx(result: dict) -> bytes:
    wb = Workbook()
    spec = result.get("spec", {})
    acc = result.get("accounting", {})
    spec_label = (f"Ø{spec.get('od')}×{spec.get('thickness')}"
                  if spec.get("profile_type") == "tube" else f"Ø{spec.get('od')}")

    ws = wb.active
    ws.title = "汇总"
    rows = [
        ("项目", "数值"),
        ("材料牌号", spec.get("material", "-")),
        ("料型", "管材" if spec.get("profile_type") == "tube" else "棒材"),
        ("规格", spec_label),
        ("总利用率", f"{result['overall_utilization'] * 100:.1f}%"),
        ("用新料(根)", result["bars_used"]),
        ("用余料(根)", result.get("remnants_used", 0)),
        ("缺口(件)", result["total_shortfall"]),
        ("锯缝(mm)", result.get("kerf", 0)),
    ]
    if acc:
        rows += [
            ("截面积(mm²)", acc.get("section_area_mm2", "-")),
            ("每米重量(kg/m)", acc.get("weight_kg_per_m", "-")),
            ("净产品重量(kg)", acc.get("product_weight_kg", "-")),
            ("投入新料重量(kg)", acc.get("input_weight_kg", "-")),
            ("可回收余料(kg)", acc.get("remnant_weight_kg", "-")),
        ]
    for r in rows:
        ws.append(r)
    _style_header(ws, 1, 2)
    _autofit(ws, [22, 22])

    ws2 = wb.create_sheet("切割单")
    ws2.append(["料号", "类型", "切割明细(长×数)", "利用率", "余料(mm)"])
    _style_header(ws2, 1, 5)
    for b in result["bars"]:
        cuts = ", ".join(f"{k}×{v}" for k, v in b["cuts"].items())
        ws2.append([b["bar_code"], "余料" if b["is_remnant"] else "新料", cuts,
                    f"{b['utilization'] * 100:.1f}%", b["remnant_length"]])
        if b["is_remnant"]:
            for c in range(1, 6):
                ws2.cell(row=ws2.max_row, column=c).fill = _REMFILL
    _autofit(ws2, [14, 8, 40, 10, 12])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------- 打印友好 HTML(浏览器导出 PDF)----------------

_CSS = """<style>
body{font-family:-apple-system,"PingFang SC",sans-serif;color:#222;margin:24px;}
h1{font-size:20px;margin:0 0 4px;} .sub{color:#666;font-size:13px;margin-bottom:14px;}
table{border-collapse:collapse;width:100%;margin-bottom:16px;font-size:13px;}
th,td{border:1px solid #999;padding:6px 8px;text-align:left;}
th{background:#4e79a7;color:#fff;}
.rem{background:#e2efda;} .kpi{display:flex;gap:24px;flex-wrap:wrap;margin-bottom:14px;}
.kpi div{font-size:13px;} .kpi b{font-size:18px;display:block;color:#4e79a7;}
@media print{body{margin:0;}}
</style>"""


def _html(title: str, sub: str, kpis: list[tuple[str, str]], tables: list[str]) -> str:
    kpi_html = "".join(f"<div><b>{v}</b>{k}</div>" for k, v in kpis)
    return (f"<!doctype html><html><head><meta charset='utf-8'><title>{title}</title>{_CSS}</head>"
            f"<body><h1>{title}</h1><div class='sub'>{sub}</div>"
            f"<div class='kpi'>{kpi_html}</div>{''.join(tables)}"
            "<script>window.onload=()=>window.print()</script></body></html>")


def sheet_html(result: dict) -> str:
    spec = result.get("spec", {})
    acc = result.get("accounting", {})
    kpis = [("总利用率", f"{result['overall_utilization'] * 100:.1f}%"),
            ("用新板", str(result["sheets_used"])),
            ("用余料", str(result.get("remnants_used", 0))),
            ("缺口", str(result["total_shortfall"]))]
    if acc:
        kpis += [("净产品(kg)", str(acc.get("product_weight_kg", "-"))),
                 ("投入新板(kg)", str(acc.get("input_weight_kg", "-")))]
    rows = "".join(
        f"<tr class='{'rem' if s['is_remnant'] else ''}'><td>{s['sheet_code']}</td>"
        f"<td>{'余料' if s['is_remnant'] else '新板'}</td><td>{_counts_text(s['part_counts'])}</td>"
        f"<td>{s['utilization'] * 100:.1f}%</td></tr>" for s in result["sheets"])
    table = ("<table><thead><tr><th>板料编码</th><th>类型</th><th>零件号×数量</th><th>利用率</th></tr></thead>"
             f"<tbody>{rows}</tbody></table>")
    sub = f"牌号 {spec.get('material', '-')} · 板厚 {spec.get('thickness', '-')}mm"
    return _html("板材出料表", sub, kpis, [table])


def bar_html(result: dict) -> str:
    spec = result.get("spec", {})
    acc = result.get("accounting", {})
    spec_label = (f"Ø{spec.get('od')}×{spec.get('thickness')}"
                  if spec.get("profile_type") == "tube" else f"Ø{spec.get('od')}")
    kpis = [("总利用率", f"{result['overall_utilization'] * 100:.1f}%"),
            ("用新料", str(result["bars_used"])),
            ("用余料", str(result.get("remnants_used", 0))),
            ("缺口", str(result["total_shortfall"]))]
    if acc:
        kpis += [("净产品(kg)", str(acc.get("product_weight_kg", "-"))),
                 ("可回收余料(kg)", str(acc.get("remnant_weight_kg", "-")))]
    rows = "".join(
        f"<tr class='{'rem' if b['is_remnant'] else ''}'><td>{b['bar_code']}</td>"
        f"<td>{'余料' if b['is_remnant'] else '新料'}</td>"
        f"<td>{', '.join(f'{k}×{v}' for k, v in b['cuts'].items())}</td>"
        f"<td>{b['utilization'] * 100:.1f}%</td><td>{b['remnant_length']}mm</td></tr>"
        for b in result["bars"])
    table = ("<table><thead><tr><th>料号</th><th>类型</th><th>切割明细</th><th>利用率</th><th>余料</th></tr></thead>"
             f"<tbody>{rows}</tbody></table>")
    sub = f"牌号 {spec.get('material', '-')} · {'管材' if spec.get('profile_type') == 'tube' else '棒材'} {spec_label} · 锯缝 {result.get('kerf', 0)}mm"
    return _html("型材切割单", sub, kpis, [table])
